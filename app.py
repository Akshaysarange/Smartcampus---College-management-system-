import os
import secrets
import math

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_mysqldb import MySQL
from datetime import datetime, timedelta
from openpyxl import Workbook,load_workbook 


app = Flask(__name__)
app.secret_key = 'smartcampus_secret_key'

# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'root'
app.config['MYSQL_DB'] = 'smartcampus'

mysql = MySQL(app)


# =========================
# Helper functions
# =========================
def is_logged_in(role=None):
    if 'user_id' not in session:
        return False
    if role and session.get('role') != role:
        return False
    return True


def next_username(role, start_number):
    """Generate next email-style username for teacher/student."""
    cur = mysql.connection.cursor()
    cur.execute("SELECT username FROM users WHERE role = %s ORDER BY id DESC LIMIT 1", (role,))
    last = cur.fetchone()
    cur.close()

    if last:
        last_id = int(last[0].split('@')[0])
        new_id = last_id + 1
    else:
        new_id = start_number

    return f"{new_id}@college.ac.in"


def get_attendance_summary(student_id, year_name=None):
    """Return present/absent/leave/total/percent for a student, optional year filter."""
    cur = mysql.connection.cursor()

    if year_name:
        cur.execute("""
            SELECT a.status, COUNT(*)
            FROM attendance a
            JOIN subjects s ON a.subject_id = s.id
            JOIN years y ON s.year_id = y.id
            WHERE a.student_id = %s AND y.name = %s
            GROUP BY a.status
        """, (student_id, year_name))
    else:
        cur.execute("""
            SELECT status, COUNT(*)
            FROM attendance
            WHERE student_id = %s
            GROUP BY status
        """, (student_id,))

    rows = cur.fetchall()
    cur.close()

    present = absent = leave = 0

    for status, count in rows:
        if status == 'P':
            present = count
        elif status == 'A':
            absent = count
        elif status == 'L':
            leave = count

    total = present + absent + leave
    percent = round((present / total) * 100, 2) if total > 0 else 0

    return {
        'present': present,
        'absent': absent,
        'leave': leave,
        'total': total,
        'percent': percent
    }


# =========================
# LOGIN ROUTES
# =========================
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * FROM users WHERE BINARY username = %s AND BINARY password = %s",
            (username, password)
        )
        user = cur.fetchone()
        cur.close()

        if user:
            session['user_id'] = user[0]
            session['username'] = user[1]
            session['role'] = user[3]
            session['first_login'] = user[4]

            if user[4] == 1:
                return redirect(url_for('change_password'))

            if user[3] == 'admin':
                return redirect(url_for('admin_find'))
            if user[3] == 'teacher':
                return redirect(url_for('teacher_find'))
            if user[3] == 'student':
                return redirect(url_for('student_dashboard'))
        else:
            error = 'Invalid username or password!'

    return render_template('login.html', error=error)


@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    """First-login password change/skip page for all roles."""
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        action = request.form.get('action')

        cur = mysql.connection.cursor()

        if action == 'skip':
            cur.execute("UPDATE users SET first_login = 0 WHERE id = %s", (session['user_id'],))
            session['first_login'] = 0

        elif action == 'change':
            new_password = request.form['new_password']
            cur.execute(
                "UPDATE users SET password = %s, first_login = 0 WHERE id = %s",
                (new_password, session['user_id'])
            )
            session['first_login'] = 0

        mysql.connection.commit()
        cur.close()

        if session['role'] == 'admin':
            return redirect(url_for('admin_find'))
        if session['role'] == 'teacher':
            return redirect(url_for('teacher_find'))
        if session['role'] == 'student':
            return redirect(url_for('student_dashboard'))

    return render_template('change_password.html')


# =========================
# ADMIN ROUTES
# =========================
@app.route('/admin/find')
def admin_find():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))
    return render_template('admin/find.html')


@app.route('/admin/find-data/<type_name>/<dept_name>/<year_name>')
def admin_find_data(type_name, dept_name, year_name):
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    if type_name == 'teacher':
        cur.execute("""
            SELECT t.name, u.phone, t.username, u.password
            FROM teachers t
            JOIN users u ON t.user_id = u.id
            JOIN departments d ON t.dept_id = d.id
            WHERE d.name = %s
            ORDER BY t.id
        """, (dept_name,))

        rows = cur.fetchall()

        result = [
            {
                'name': r[0],
                'phone': r[1],
                'username': r[2],
                'password': r[3]
            }
            for r in rows
        ]

    else:
        cur.execute("""
            SELECT st.name, st.roll_no, st.username, u.password
            FROM students st
            JOIN users u ON st.user_id = u.id
            JOIN departments d ON st.dept_id = d.id
            JOIN years y ON st.year_id = y.id
            WHERE d.name = %s AND y.name = %s
            ORDER BY st.id
        """, (dept_name, year_name))

        rows = cur.fetchall()
        result = [
            {'name': r[0], 'roll': r[1], 'username': r[2], 'password': r[3]}
            for r in rows
        ]

    cur.close()
    return jsonify(result)

@app.route('/admin/search/<keyword>')
def admin_search(keyword):
    if not is_logged_in('admin'):
        return jsonify({
            "success": False,
            "message": "Unauthorized access."
        }), 401

    cur = None

    try:
        cur = mysql.connection.cursor()
        key = f"%{keyword}%"

        # Student records
        cur.execute("""
            SELECT
                'Student' AS record_type,
                st.name,
                d.name AS department,
                y.name AS year_name,
                st.roll_no,
                st.username,
                u.phone,
                u.password,
                '' AS fy_subjects,
                '' AS sy_subjects,
                '' AS ty_subjects
            FROM students st
            JOIN users u
                ON u.id = st.user_id
            JOIN departments d
                ON d.id = st.dept_id
            JOIN years y
                ON y.id = st.year_id
            WHERE st.name LIKE %s
               OR st.roll_no LIKE %s
               OR st.username LIKE %s
               OR u.phone LIKE %s
        """, (
            key,
            key,
            key,
            key
        ))

        student_rows = cur.fetchall()

        # Teacher records with year-wise assigned subjects
        cur.execute("""
            SELECT
                'Teacher' AS record_type,
                t.name,
                d.name AS department,
                '-' AS year_name,
                '-' AS roll_no,
                t.username,
                u.phone,
                u.password,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 1 THEN s.name
                    END
                    ORDER BY s.name
                    SEPARATOR ', '
                ) AS fy_subjects,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 2 THEN s.name
                    END
                    ORDER BY s.name
                    SEPARATOR ', '
                ) AS sy_subjects,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 3 THEN s.name
                    END
                    ORDER BY s.name
                    SEPARATOR ', '
                ) AS ty_subjects

            FROM teachers t

            JOIN users u
                ON u.id = t.user_id

            JOIN departments d
                ON d.id = t.dept_id

            LEFT JOIN teacher_subjects ts
                ON ts.teacher_id = t.id

            LEFT JOIN subjects s
                ON s.id = ts.subject_id

            WHERE t.name LIKE %s
               OR t.username LIKE %s
               OR u.phone LIKE %s
               OR s.name LIKE %s

            GROUP BY
                t.id,
                t.name,
                d.name,
                t.username,
                u.phone,
                u.password

            ORDER BY t.id
        """, (
            key,
            key,
            key,
            key
        ))

        teacher_rows = cur.fetchall()

        rows = student_rows + teacher_rows

        result = []

        for row in rows:
            result.append({
                "type": row[0],
                "name": row[1],
                "department": row[2],
                "year": row[3],
                "roll": row[4],
                "username": row[5],
                "phone": row[6],
                "password": row[7],
                "fy_subjects": row[8] or "",
                "sy_subjects": row[9] or "",
                "ty_subjects": row[10] or ""
            })

        return jsonify(result), 200

    except Exception as error:
        print("Admin search error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to search records."
        }), 500

    finally:
        if cur:
            cur.close()

@app.route('/admin/subjects/<int:dept_id>/<int:year_id>')
def admin_subjects_by_dept_year(dept_id, year_id):
    """Use this route in JS to load subjects dynamically instead of hardcoded data."""
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id, name
        FROM subjects
        WHERE dept_id = %s AND year_id = %s
        ORDER BY id
    """, (dept_id, year_id))
    subjects = cur.fetchall()
    cur.close()

    return jsonify([
        {'id': s[0], 'name': s[1]}
        for s in subjects
    ])


@app.route('/admin/subjects/list/<int:dept_id>/<int:year_id>')
def admin_subjects_list(dept_id, year_id):
    if not is_logged_in('admin'):
        return jsonify({
            'success': False,
            'message': 'Unauthorized'
        }), 401

    cur = mysql.connection.cursor()

    try:
        cur.execute("""
            SELECT id, name
            FROM subjects
            WHERE dept_id = %s
              AND year_id = %s
            ORDER BY id
        """, (dept_id, year_id))

        subjects = cur.fetchall()

        result = []

        for subject in subjects:
            result.append({
                'id': subject[0],
                'name': subject[1]
            })

        return jsonify(result)

    except Exception as error:
        print("Subject loading error:", error)

        return jsonify({
            'success': False,
            'message': 'Unable to load subjects'
        }), 500

    finally:
        cur.close()

@app.route('/admin/teachers')
def admin_teachers():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    cur = None

    try:
        cur = mysql.connection.cursor()

        cur.execute("""
            SELECT
                COUNT(*) AS total_teachers,
                SUM(CASE WHEN dept_id = 1 THEN 1 ELSE 0 END) AS cs_teachers,
                SUM(CASE WHEN dept_id = 2 THEN 1 ELSE 0 END) AS it_teachers,
                SUM(CASE WHEN dept_id = 3 THEN 1 ELSE 0 END) AS dsda_teachers,
                SUM(CASE WHEN dept_id = 4 THEN 1 ELSE 0 END) AS ai_teachers
            FROM teachers
        """)

        counts = cur.fetchone()

        total_teachers = counts[0] or 0
        cs_teachers = counts[1] or 0
        it_teachers = counts[2] or 0
        dsda_teachers = counts[3] or 0
        ai_teachers = counts[4] or 0

        return render_template(
            'admin/teachers.html',
            total_teachers=total_teachers,
            cs_teachers=cs_teachers,
            it_teachers=it_teachers,
            dsda_teachers=dsda_teachers,
            ai_teachers=ai_teachers
        )

    except Exception as error:
        print("Admin teachers page error:", error)

        flash(
            "Unable to load teacher page. Please try again.",
            "error"
        )

        return redirect(url_for('admin_find'))

    finally:
        if cur:
            cur.close()

@app.route('/admin/teachers/add', methods=['POST'])
def admin_teachers_add():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    dept_id = request.form.get('dept_id', '').strip()

    fy_subject_ids = request.form.getlist('fy_subject_ids')
    sy_subject_ids = request.form.getlist('sy_subject_ids')
    ty_subject_ids = request.form.getlist('ty_subject_ids')

    # Empty values aur duplicates remove karo
    fy_subject_ids = list(dict.fromkeys(
        subject_id
        for subject_id in fy_subject_ids
        if subject_id
    ))

    sy_subject_ids = list(dict.fromkeys(
        subject_id
        for subject_id in sy_subject_ids
        if subject_id
    ))

    ty_subject_ids = list(dict.fromkeys(
        subject_id
        for subject_id in ty_subject_ids
        if subject_id
    ))

    # Basic details validation
    if len(name) < 2:
        flash(
            "Please enter a valid teacher name.",
            "error"
        )
        return redirect(url_for('admin_teachers'))

    if not phone.isdigit() or len(phone) != 10:
        flash(
            "Please enter a valid 10-digit phone number.",
            "error"
        )
        return redirect(url_for('admin_teachers'))

    if not dept_id.isdigit():
        flash(
            "Please select a valid department.",
            "error"
        )
        return redirect(url_for('admin_teachers'))

    # Har year minimum 1 aur maximum 6 subjects
    if not 1 <= len(fy_subject_ids) <= 6:
        flash(
            "Please select between 1 and 6 FY subjects.",
            "error"
        )
        return redirect(url_for('admin_teachers'))

    if not 1 <= len(sy_subject_ids) <= 6:
        flash(
            "Please select between 1 and 6 SY subjects.",
            "error"
        )
        return redirect(url_for('admin_teachers'))

    if not 1 <= len(ty_subject_ids) <= 6:
        flash(
            "Please select between 1 and 6 TY subjects.",
            "error"
        )
        return redirect(url_for('admin_teachers'))

    cur = None

    try:
        cur = mysql.connection.cursor()

        def validate_subjects(
            subject_ids,
            year_id,
            year_name
        ):
            placeholders = ", ".join(
                ["%s"] * len(subject_ids)
            )

            query = f"""
                SELECT id
                FROM subjects
                WHERE dept_id = %s
                  AND year_id = %s
                  AND id IN ({placeholders})
            """

            parameters = (
                [dept_id, year_id]
                + subject_ids
            )

            cur.execute(
                query,
                parameters
            )

            valid_subject_ids = {
                str(row[0])
                for row in cur.fetchall()
            }

            if valid_subject_ids != set(subject_ids):
                flash(
                    f"One or more selected {year_name} subjects are invalid.",
                    "error"
                )
                return False

            return True

        if not validate_subjects(
            fy_subject_ids,
            1,
            "FY"
        ):
            return redirect(
                url_for('admin_teachers')
            )

        if not validate_subjects(
            sy_subject_ids,
            2,
            "SY"
        ):
            return redirect(
                url_for('admin_teachers')
            )

        if not validate_subjects(
            ty_subject_ids,
            3,
            "TY"
        ):
            return redirect(
                url_for('admin_teachers')
            )

        username = next_username(
            'teacher',
            1000001
        )

        password = "teacher"

        cur.execute("""
            INSERT INTO users
            (
                username,
                password,
                role,
                first_login,
                phone
            )
            VALUES
            (
                %s,
                %s,
                'teacher',
                1,
                %s
            )
        """, (
            username,
            password,
            phone
        ))

        user_id = cur.lastrowid

        cur.execute("""
            INSERT INTO teachers
            (
                user_id,
                name,
                username,
                dept_id
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s
            )
        """, (
            user_id,
            name,
            username,
            dept_id
        ))

        teacher_id = cur.lastrowid

        all_subject_ids = (
            fy_subject_ids
            + sy_subject_ids
            + ty_subject_ids
        )

        subject_assignments = [
            (
                teacher_id,
                subject_id
            )
            for subject_id in all_subject_ids
        ]

        cur.executemany("""
            INSERT INTO teacher_subjects
            (
                teacher_id,
                subject_id
            )
            VALUES
            (
                %s,
                %s
            )
        """, subject_assignments)

        mysql.connection.commit()

        flash(
            f"Teacher added successfully! Username: {username}",
            "success"
        )

    except Exception as error:
        mysql.connection.rollback()

        print(
            "Teacher add error:",
            error
        )

        flash(
            "Unable to add teacher. Please try again.",
            "error"
        )

    finally:
        if cur:
            cur.close()

    return redirect(
        url_for('admin_teachers')
    )


@app.route('/admin/teachers/remove', methods=['POST'])
def admin_teachers_remove():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    teacher_id = request.form.get('teacher_id', '').strip()

    if not teacher_id:
        flash('Teacher ID is required!', 'error')
        return redirect(url_for('admin_teachers'))

    cur = mysql.connection.cursor()

    try:
        cur.execute("""
            SELECT user_id
            FROM teachers
            WHERE id = %s
        """, (teacher_id,))

        result = cur.fetchone()

        if not result:
            flash('Teacher not found!', 'error')
            return redirect(url_for('admin_teachers'))

        user_id = result[0]

        cur.execute("""
            DELETE FROM teacher_subjects
            WHERE teacher_id = %s
        """, (teacher_id,))

        cur.execute("""
            DELETE FROM teachers
            WHERE id = %s
        """, (teacher_id,))

        cur.execute("""
            DELETE FROM users
            WHERE id = %s
        """, (user_id,))

        mysql.connection.commit()

        flash('Teacher removed successfully!', 'success')

    except Exception as error:
        mysql.connection.rollback()
        print("Teacher remove error:", error)

        flash(
            'Unable to remove teacher. Please try again.',
            'error'
        )

    finally:
        cur.close()

    return redirect(url_for('admin_teachers'))


@app.route('/admin/teachers/list/<dept_name>')
def admin_teachers_list(dept_name):
    if not is_logged_in('admin'):
        return jsonify({
            "success": False,
            "message": "Unauthorized access."
        }), 401

    cur = None

    try:
        cur = mysql.connection.cursor()

        cur.execute("""
            SELECT
                t.id,
                t.name,
                t.username,
                u.password,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 1 THEN s.name
                    END
                    ORDER BY s.id
                    SEPARATOR ', '
                ) AS fy_subjects,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 2 THEN s.name
                    END
                    ORDER BY s.id
                    SEPARATOR ', '
                ) AS sy_subjects,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 3 THEN s.name
                    END
                    ORDER BY s.id
                    SEPARATOR ', '
                ) AS ty_subjects

            FROM teachers t

            JOIN users u
                ON u.id = t.user_id

            JOIN departments d
                ON d.id = t.dept_id

            LEFT JOIN teacher_subjects ts
                ON ts.teacher_id = t.id

            LEFT JOIN subjects s
                ON s.id = ts.subject_id

            WHERE d.name = %s

            GROUP BY
                t.id,
                t.name,
                t.username,
                u.password

            ORDER BY t.id ASC
        """, (dept_name,))

        rows = cur.fetchall()

        teachers = []

        for row in rows:
            teachers.append({
                "id": row[0],
                "name": row[1],
                "username": row[2],
                "password": row[3],
                "fy_subjects": row[4] or "",
                "sy_subjects": row[5] or "",
                "ty_subjects": row[6] or ""
            })

        return jsonify(teachers), 200

    except Exception as error:
        print("Teacher list error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to load teachers."
        }), 500

    finally:
        if cur:
            cur.close()

@app.route('/admin/teachers/search/<keyword>')
def admin_search_teachers(keyword):
    if not is_logged_in('admin'):
        return jsonify({
            "success": False,
            "message": "Unauthorized access."
        }), 401

    keyword = str(keyword).strip()

    if not keyword:
        return jsonify([]), 200

    search_value = f"%{keyword}%"

    cur = None

    try:
        cur = mysql.connection.cursor()

        cur.execute("""
            SELECT
                t.id,
                t.name,
                t.username,
                u.password,
                u.phone,
                d.name,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 1 THEN s.name
                    END
                    ORDER BY s.id
                    SEPARATOR ', '
                ) AS fy_subjects,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 2 THEN s.name
                    END
                    ORDER BY s.id
                    SEPARATOR ', '
                ) AS sy_subjects,

                GROUP_CONCAT(
                    DISTINCT CASE
                        WHEN s.year_id = 3 THEN s.name
                    END
                    ORDER BY s.id
                    SEPARATOR ', '
                ) AS ty_subjects

            FROM teachers t

            JOIN users u
                ON u.id = t.user_id

            JOIN departments d
                ON d.id = t.dept_id

            LEFT JOIN teacher_subjects ts
                ON ts.teacher_id = t.id

            LEFT JOIN subjects s
                ON s.id = ts.subject_id

            WHERE
                t.name LIKE %s
                OR t.username LIKE %s
                OR u.phone LIKE %s
                OR d.name LIKE %s

            GROUP BY
                t.id,
                t.name,
                t.username,
                u.password,
                u.phone,
                d.name

            ORDER BY t.name ASC
        """, (
            search_value,
            search_value,
            search_value,
            search_value
        ))

        rows = cur.fetchall()

        teachers = []

        for row in rows:
            teachers.append({
                "id": row[0],
                "name": row[1],
                "username": row[2],
                "password": row[3],
                "phone": row[4] or "",
                "department": row[5] or "",
                "fy_subjects": row[6] or "",
                "sy_subjects": row[7] or "",
                "ty_subjects": row[8] or ""
            })

        return jsonify(teachers), 200

    except Exception as error:
        print("Teacher search error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to search teachers."
        }), 500

    finally:
        if cur:
            cur.close()


@app.route('/admin/students')
def admin_students():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    cur.execute("SELECT COUNT(*) FROM students")
    total_students = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM students WHERE dept_id = 1")
    cs_students = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM students WHERE dept_id = 2")
    it_students = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM students WHERE dept_id = 3")
    dsda_students = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM students WHERE dept_id = 4")
    ai_students = cur.fetchone()[0]

    cur.close()

    return render_template(
        'admin/students.html',
        total_students=total_students,
        cs_students=cs_students,
        it_students=it_students,
        dsda_students=dsda_students,
        ai_students=ai_students
    )


@app.route('/admin/students/add', methods=['POST'])
def admin_students_add():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    name = request.form['name']
    phone = request.form['phone']
    dept_id = request.form['dept_id']
    year_id = request.form['year_id']

    username = next_username('student', 2000001)
    password = 'student'

    dept_codes = {
        '1': 'CS',
        '2': 'IT',
        '3': 'DSDA',
        '4': 'AI'
    }

    year_codes = {
        '1': 'FY',
        '2': 'SY',
        '3': 'TY'
    }

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT COUNT(*)
        FROM students
        WHERE dept_id = %s AND year_id = %s
    """, (dept_id, year_id))
    next_roll_no = cur.fetchone()[0] + 1

    roll_no = f"{dept_codes[dept_id]}-{year_codes[year_id]}-{next_roll_no:03d}"

    cur.execute(
        "INSERT INTO users (username, password, role, first_login, phone) VALUES (%s, %s, 'student', 1, %s)",
        (username, password, phone)
    )
    user_id = cur.lastrowid

    cur.execute(
        "INSERT INTO students (user_id, name, username, roll_no, dept_id, year_id) VALUES (%s, %s, %s, %s, %s, %s)",
        (user_id, name, username, roll_no, dept_id, year_id)
    )

    mysql.connection.commit()
    cur.close()

    flash(f'Student added successfully! Username: {username}', 'success')
    return redirect(url_for('admin_students'))


@app.route('/admin/students/list/<dept_name>/<year_name>')
def admin_students_list(dept_name, year_name):
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT st.id, st.name, st.roll_no, st.username
        FROM students st
        JOIN departments d ON st.dept_id = d.id
        JOIN years y ON st.year_id = y.id
        WHERE d.name = %s AND y.name = %s
        ORDER BY st.id
    """, (dept_name, year_name))

    students = cur.fetchall()
    cur.close()

    result = [
        {'id': s[0], 'name': s[1], 'roll': s[2], 'username': s[3]}
        for s in students
    ]

    return jsonify(result)

@app.route('/admin/students/search/<keyword>')
def admin_search_students(keyword):
    if not is_logged_in('admin'):
        return jsonify({
            "success": False,
            "message": "Unauthorized access."
        }), 401

    keyword = str(keyword).strip()

    if not keyword:
        return jsonify([]), 200

    search_value = f"%{keyword}%"

    cur = None

    try:
        cur = mysql.connection.cursor()

        cur.execute("""
            SELECT
                st.id,
                st.name,
                st.roll_no,
                st.username,
                u.phone,
                d.name,
                y.name
            FROM students st

            JOIN users u
                ON u.id = st.user_id

            JOIN departments d
                ON d.id = st.dept_id

            JOIN years y
                ON y.id = st.year_id

            WHERE
                st.name LIKE %s
                OR st.roll_no LIKE %s
                OR st.username LIKE %s
                OR u.phone LIKE %s
                OR d.name LIKE %s
                OR y.name LIKE %s

            ORDER BY st.name ASC
        """, (
            search_value,
            search_value,
            search_value,
            search_value,
            search_value,
            search_value
        ))

        rows = cur.fetchall()

        students = []

        for row in rows:
            students.append({
                "id": row[0],
                "name": row[1],
                "roll": row[2],
                "username": row[3],
                "phone": row[4] or "",
                "department": row[5] or "",
                "year": row[6] or ""
            })

        return jsonify(students), 200

    except Exception as error:
        print("Student search error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to search students."
        }), 500

    finally:
        if cur:
            cur.close()

@app.route('/admin/students/remove', methods=['POST'])
def admin_students_remove():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    student_id = request.form['student_id']

    cur = mysql.connection.cursor()
    cur.execute("SELECT user_id FROM students WHERE id = %s", (student_id,))
    result = cur.fetchone()

    if result:
        user_id = result[0]

        cur.execute("DELETE FROM attendance WHERE student_id = %s", (student_id,))
        cur.execute("DELETE FROM marks WHERE student_id = %s", (student_id,))
        cur.execute("DELETE FROM students WHERE id = %s", (student_id,))
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))

        mysql.connection.commit()
        flash('Student removed successfully!', 'success')
    else:
        flash('Student not found!', 'error')

    cur.close()
    return redirect(url_for('admin_students'))


@app.route('/admin/profile')
def admin_profile():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()
    cur.execute(
        "SELECT id, username, password, phone FROM users WHERE id = %s",
        (session['user_id'],)
    )
    admin = cur.fetchone()
    cur.close()

    hour = datetime.now().hour

    if hour < 12:
        greeting = 'Good Morning'
    elif hour < 17:
        greeting = 'Good Afternoon'
    else:
        greeting = 'Good Evening'

    return render_template(
        'admin/profile.html',
        greeting=greeting,
        name=admin[1],
        admin_id=admin[0],
        username=admin[1],
        password=admin[2],
        phone=admin[3] or 'Not Available'
    )


@app.route('/admin/change-password', methods=['GET', 'POST'])
def admin_change_password():
    if not is_logged_in('admin'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        old_password = request.form['old_password']
        new_password = request.form['new_password']

        cur = mysql.connection.cursor()
        cur.execute("SELECT password FROM users WHERE id = %s", (session['user_id'],))
        user = cur.fetchone()

        if user and str(user[0]) == str(old_password):
            cur.execute(
                "UPDATE users SET password = %s WHERE id = %s",
                (new_password, session['user_id'])
            )
            mysql.connection.commit()
            flash('Password changed successfully!', 'success')
        else:
            flash('Old password is incorrect!', 'error')

        cur.close()
        return redirect(url_for('admin_change_password'))

    return render_template('admin/password.html')


# =========================
# TEACHER ROUTES
# =========================
@app.route('/teacher/find')
def teacher_find():
    if not is_logged_in('teacher'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    # Teacher Department
    cur.execute("""
        SELECT t.dept_id, d.name, t.name
        FROM teachers t
        JOIN departments d ON t.dept_id = d.id
        WHERE t.user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()

    dept_id = teacher[0]
    dept_name = teacher[1]
    teacher_name = teacher[2]

    # FY Count
    cur.execute("""
        SELECT COUNT(*)
        FROM students
        WHERE dept_id=%s AND year_id=1
    """, (dept_id,))
    fy_count = cur.fetchone()[0]

    # SY Count
    cur.execute("""
        SELECT COUNT(*)
        FROM students
        WHERE dept_id=%s AND year_id=2
    """, (dept_id,))
    sy_count = cur.fetchone()[0]

    # TY Count
    cur.execute("""
        SELECT COUNT(*)
        FROM students
        WHERE dept_id=%s AND year_id=3
    """, (dept_id,))
    ty_count = cur.fetchone()[0]

    cur.close()

    return render_template(
        "teacher/find.html",
        teacher_name=teacher_name,
        department=dept_name,
        fy_count=fy_count,
        sy_count=sy_count,
        ty_count=ty_count
    )

@app.route('/teacher/subjects/<year_name>')
def teacher_subjects(year_name):
    if not is_logged_in('teacher'):
        return jsonify([])

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT t.id, t.dept_id
        FROM teachers t
        WHERE t.user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()

    if not teacher:
        cur.close()
        return jsonify([])

    teacher_id = teacher[0]
    dept_id = teacher[1]

    cur.execute("""
        SELECT s.id, s.name
        FROM subjects s
        JOIN teacher_subjects ts ON s.id = ts.subject_id
        JOIN years y ON s.year_id = y.id
        WHERE ts.teacher_id = %s
        AND s.dept_id = %s
        AND y.name = %s
        ORDER BY s.id
    """, (teacher_id, dept_id, year_name))

    rows = cur.fetchall()
    cur.close()

    return jsonify([
        {"id": r[0], "name": r[1]}
        for r in rows
    ])

@app.route('/teacher/attendance/students/<year_name>/<int:subject_id>/<attendance_date>')
def attendance_students(year_name, subject_id, attendance_date):
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    # Invalid ya future date ko block karo
    try:
        selected_date = datetime.strptime(
            attendance_date,
            "%Y-%m-%d"
        ).date()
    except (ValueError, TypeError):
        return jsonify({
            "success": False,
            "message": "Invalid date format"
        }), 400

    if selected_date > datetime.now().date():
        return jsonify({
            "success": False,
            "message": "Future date attendance is not allowed"
        }), 400

    cur = mysql.connection.cursor()

    try:
        # Logged-in teacher ki department nikalo
        cur.execute("""
            SELECT id, dept_id
            FROM teachers
            WHERE user_id = %s
        """, (session['user_id'],))

        teacher = cur.fetchone()

        if not teacher:
            return jsonify({
                "success": False,
                "message": "Teacher not found"
            }), 404

        teacher_id = teacher[0]
        dept_id = teacher[1]

        # Check karo selected subject teacher ko assigned hai ya nahi
        cur.execute("""
            SELECT s.id
            FROM subjects s
            JOIN teacher_subjects ts
                ON ts.subject_id = s.id
            JOIN years y
                ON s.year_id = y.id
            WHERE ts.teacher_id = %s
              AND s.id = %s
              AND s.dept_id = %s
              AND y.name = %s
        """, (
            teacher_id,
            subject_id,
            dept_id,
            year_name
        ))

        assigned_subject = cur.fetchone()

        if not assigned_subject:
            return jsonify({
                "success": False,
                "message": "This subject is not assigned to you"
            }), 403

        # Students aur existing attendance load karo
        cur.execute("""
            SELECT
                st.id,
                st.name,
                st.roll_no,
                IFNULL(a.status, 'P') AS status
            FROM students st
            JOIN years y
                ON st.year_id = y.id
            LEFT JOIN attendance a
                ON a.student_id = st.id
                AND a.subject_id = %s
                AND a.date = %s
            WHERE st.dept_id = %s
              AND y.name = %s
            ORDER BY st.roll_no
        """, (
            subject_id,
            attendance_date,
            dept_id,
            year_name
        ))

        rows = cur.fetchall()

        return jsonify([
            {
                "id": row[0],
                "name": row[1],
                "roll": row[2],
                "status": row[3]
            }
            for row in rows
        ])

    except Exception as error:
        print("Attendance students error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to load students"
        }), 500

    finally:
        cur.close()

@app.route('/teacher/attendance/generate-otp', methods=['POST'])
def generate_attendance_otp():
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Invalid request data."
        }), 400

    year_name = str(data.get("year", "")).strip().upper()
    subject_id = data.get("subject_id")
    latitude = data.get("latitude")
    longitude = data.get("longitude")
    allowed_radius = data.get("allowed_radius", 100)

    if year_name not in ("FY", "SY", "TY"):
        return jsonify({
            "success": False,
            "message": "Please select a valid year."
        }), 400

    try:
        subject_id = int(subject_id)
        latitude = float(latitude)
        longitude = float(longitude)
        allowed_radius = int(allowed_radius)

    except (TypeError, ValueError):
        return jsonify({
            "success": False,
            "message": "Invalid subject, location or radius."
        }), 400

    if latitude < -90 or latitude > 90:
        return jsonify({
            "success": False,
            "message": "Invalid latitude."
        }), 400

    if longitude < -180 or longitude > 180:
        return jsonify({
            "success": False,
            "message": "Invalid longitude."
        }), 400

    if allowed_radius not in (50, 100, 150, 200):
        return jsonify({
            "success": False,
            "message": "Invalid attendance radius."
        }), 400

    cur = mysql.connection.cursor()

    try:
        # Logged-in teacher details
        cur.execute("""
            SELECT
                t.id,
                t.dept_id
            FROM teachers t
            WHERE t.user_id = %s
        """, (session['user_id'],))

        teacher = cur.fetchone()

        if not teacher:
            return jsonify({
                "success": False,
                "message": "Teacher not found."
            }), 404

        teacher_id = teacher[0]
        teacher_dept_id = teacher[1]

        # Selected year details
        cur.execute("""
            SELECT id
            FROM years
            WHERE name = %s
        """, (year_name,))

        year_row = cur.fetchone()

        if not year_row:
            return jsonify({
                "success": False,
                "message": "Selected year was not found."
            }), 404

        year_id = year_row[0]

        # Subject teacher ko assigned hai ya nahi
        cur.execute("""
            SELECT
                s.id,
                s.name
            FROM subjects s
            JOIN teacher_subjects ts
                ON ts.subject_id = s.id
            WHERE ts.teacher_id = %s
              AND s.id = %s
              AND s.dept_id = %s
              AND s.year_id = %s
        """, (
            teacher_id,
            subject_id,
            teacher_dept_id,
            year_id
        ))

        subject = cur.fetchone()

        if not subject:
            return jsonify({
                "success": False,
                "message": (
                    "This subject is not assigned to you "
                    "for the selected year."
                )
            }), 403

        subject_name = subject[1]

        # Teacher ke previous active OTP sessions close karo
        cur.execute("""
            UPDATE attendance_otp_sessions
            SET is_active = 0
            WHERE teacher_id = %s
              AND is_active = 1
        """, (teacher_id,))

        # Secure 6-digit OTP
        otp_code = str(
            secrets.randbelow(900000) + 100000
        )

        created_at = datetime.now()
        expires_at = created_at + timedelta(minutes=5)
        attendance_date = created_at.date()

        # New OTP session insert
        cur.execute("""
            INSERT INTO attendance_otp_sessions (
                teacher_id,
                subject_id,
                year_id,
                otp_code,
                teacher_latitude,
                teacher_longitude,
                allowed_radius,
                expires_at,
                is_active
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, 1
            )
        """, (
            teacher_id,
            subject_id,
            year_id,
            otp_code,
            latitude,
            longitude,
            allowed_radius,
            expires_at
        ))

        otp_session_id = cur.lastrowid

        # Selected department + year ke sab students nikalo
        cur.execute("""
            SELECT st.id
            FROM students st
            WHERE st.dept_id = %s
              AND st.year_id = %s
            ORDER BY st.id
        """, (
            teacher_dept_id,
            year_id
        ))

        student_rows = cur.fetchall()

        # Jinki attendance row nahi hai unko default Absent mark karo
        # Existing Present ko dobara Absent nahi karega
        for student_row in student_rows:
            student_id = student_row[0]

            cur.execute("""
                INSERT IGNORE INTO attendance (
                    student_id,
                    subject_id,
                    date,
                    status
                )
                VALUES (%s, %s, %s, 'A')
            """, (
                student_id,
                subject_id,
                attendance_date
            ))

        # Excel ke liye complete attendance list
        cur.execute("""
            SELECT
                st.name,
                st.roll_no,
                a.status
            FROM attendance a
            JOIN students st
                ON st.id = a.student_id
            WHERE a.subject_id = %s
              AND a.date = %s
            ORDER BY st.roll_no
        """, (
            subject_id,
            attendance_date
        ))

        attendance_rows = cur.fetchall()

        # Subject information for Excel
        cur.execute("""
            SELECT
                d.name,
                y.name,
                s.name
            FROM subjects s
            JOIN departments d
                ON d.id = s.dept_id
            JOIN years y
                ON y.id = s.year_id
            WHERE s.id = %s
        """, (subject_id,))

        subject_info = cur.fetchone()

        if not subject_info:
            mysql.connection.rollback()

            return jsonify({
                "success": False,
                "message": "Subject information not found."
            }), 404

        mysql.connection.commit()

        department_name = subject_info[0]
        excel_year_name = subject_info[1]
        excel_subject_name = subject_info[2]

        # OTP generate hote hi Excel me sab students default A save
        update_attendance_excel(
            department_name,
            excel_year_name,
            excel_subject_name,
            attendance_date.strftime("%Y-%m-%d"),
            attendance_rows
        )

        return jsonify({
            "success": True,
            "message": "Attendance OTP generated successfully.",
            "session_id": otp_session_id,
            "otp_code": otp_code,
            "year": year_name,
            "subject_id": subject_id,
            "subject_name": subject_name,
            "allowed_radius": allowed_radius,
            "expires_at": expires_at.strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            "expires_in_seconds": 300,
            "total_students": len(student_rows)
        })

    except Exception as error:
        mysql.connection.rollback()

        print("Generate attendance OTP error:", error)

        return jsonify({
            "success": False,
            "message": (
                "Unable to generate OTP. "
                "Please try again."
            )
        }), 500

    finally:
        cur.close()

@app.route('/teacher/attendance/stop-otp', methods=['POST'])
def stop_attendance_otp():
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Invalid request."
        }), 400

    session_id = data.get("session_id")

    try:
        session_id = int(session_id)
    except (TypeError, ValueError):
        return jsonify({
            "success": False,
            "message": "Valid OTP session ID is required."
        }), 400

    cur = mysql.connection.cursor()

    try:
        # Logged-in teacher
        cur.execute("""
            SELECT
                id,
                dept_id
            FROM teachers
            WHERE user_id = %s
        """, (session['user_id'],))

        teacher = cur.fetchone()

        if not teacher:
            return jsonify({
                "success": False,
                "message": "Teacher not found."
            }), 404

        teacher_id = teacher[0]
        teacher_dept_id = teacher[1]

        # OTP session details
        cur.execute("""
            SELECT
                aos.subject_id,
                aos.year_id,
                DATE(aos.created_at),
                s.name,
                y.name,
                d.name
            FROM attendance_otp_sessions aos
            JOIN subjects s
                ON s.id = aos.subject_id
            JOIN years y
                ON y.id = aos.year_id
            JOIN departments d
                ON d.id = s.dept_id
            WHERE aos.id = %s
              AND aos.teacher_id = %s
            LIMIT 1
        """, (
            session_id,
            teacher_id
        ))

        otp_session = cur.fetchone()

        if not otp_session:
            return jsonify({
                "success": False,
                "message": "OTP session not found."
            }), 404

        subject_id = otp_session[0]
        year_id = otp_session[1]
        attendance_date = otp_session[2]
        subject_name = otp_session[3]
        year_name = otp_session[4]
        department_name = otp_session[5]

        # OTP session close karo
        cur.execute("""
            UPDATE attendance_otp_sessions
            SET is_active = 0
            WHERE id = %s
              AND teacher_id = %s
        """, (
            session_id,
            teacher_id
        ))

        # Present aur absent count
        cur.execute("""
            SELECT
                SUM(
                    CASE
                        WHEN a.status = 'P' THEN 1
                        ELSE 0
                    END
                ) AS present_count,

                SUM(
                    CASE
                        WHEN a.status = 'A' THEN 1
                        ELSE 0
                    END
                ) AS absent_count

            FROM students st

            LEFT JOIN attendance a
                ON a.student_id = st.id
                AND a.subject_id = %s
                AND a.date = %s

            WHERE st.dept_id = %s
              AND st.year_id = %s
        """, (
            subject_id,
            attendance_date,
            teacher_dept_id,
            year_id
        ))

        counts = cur.fetchone()

        present_count = int(counts[0] or 0)
        absent_count = int(counts[1] or 0)

        # Final attendance rows for Excel
        cur.execute("""
            SELECT
                st.name,
                st.roll_no,
                COALESCE(a.status, 'A') AS status
            FROM students st

            LEFT JOIN attendance a
                ON a.student_id = st.id
                AND a.subject_id = %s
                AND a.date = %s

            WHERE st.dept_id = %s
              AND st.year_id = %s

            ORDER BY st.roll_no
        """, (
            subject_id,
            attendance_date,
            teacher_dept_id,
            year_id
        ))

        attendance_rows = cur.fetchall()

        mysql.connection.commit()

        update_attendance_excel(
            department_name,
            year_name,
            subject_name,
            attendance_date.strftime("%Y-%m-%d"),
            attendance_rows
        )

        return jsonify({
            "success": True,
            "message": "OTP attendance completed successfully.",
            "present": present_count,
            "absent": absent_count,
            "total": present_count + absent_count,
            "subject": subject_name,
            "year": year_name
        })

    except Exception as error:
        mysql.connection.rollback()

        print("Stop OTP error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to complete OTP attendance."
        }), 500

    finally:
        cur.close()

@app.route('/student/attendance/verify-otp', methods=['POST'])
def verify_student_attendance_otp():
    if not is_logged_in('student'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Invalid request data."
        }), 400

    otp_code = str(data.get("otp_code", "")).strip()
    latitude = data.get("latitude")
    longitude = data.get("longitude")

    if not otp_code.isdigit() or len(otp_code) != 6:
        return jsonify({
            "success": False,
            "message": "Please enter a valid 6-digit OTP."
        }), 400

    try:
        student_latitude = float(latitude)
        student_longitude = float(longitude)
    except (TypeError, ValueError):
        return jsonify({
            "success": False,
            "message": "Invalid student location."
        }), 400

    if not -90 <= student_latitude <= 90:
        return jsonify({
            "success": False,
            "message": "Invalid latitude."
        }), 400

    if not -180 <= student_longitude <= 180:
        return jsonify({
            "success": False,
            "message": "Invalid longitude."
        }), 400

    cur = mysql.connection.cursor()

    try:
        # Logged-in student details
        cur.execute("""
            SELECT
                st.id,
                st.dept_id,
                st.year_id,
                st.name
            FROM students st
            WHERE st.user_id = %s
        """, (session['user_id'],))

        student = cur.fetchone()

        if not student:
            return jsonify({
                "success": False,
                "message": "Student not found."
            }), 404

        student_id = student[0]
        student_dept_id = student[1]
        student_year_id = student[2]

        # Expired OTP sessions deactivate karo
        cur.execute("""
            UPDATE attendance_otp_sessions
            SET is_active = 0
            WHERE is_active = 1
              AND expires_at <= NOW()
        """)

        # Active OTP session find karo
        cur.execute("""
            SELECT
                aos.id,
                aos.teacher_id,
                aos.subject_id,
                aos.year_id,
                aos.teacher_latitude,
                aos.teacher_longitude,
                aos.allowed_radius,
                aos.expires_at,
                s.name,
                s.dept_id,
                y.name
            FROM attendance_otp_sessions aos
            JOIN subjects s
                ON s.id = aos.subject_id
            JOIN years y
                ON y.id = aos.year_id
            WHERE aos.otp_code = %s
              AND aos.is_active = 1
              AND aos.expires_at > NOW()
            ORDER BY aos.id DESC
            LIMIT 1
        """, (otp_code,))

        otp_session = cur.fetchone()

        if not otp_session:
            mysql.connection.commit()

            return jsonify({
                "success": False,
                "message": "OTP is invalid or expired."
            }), 400

        otp_session_id = otp_session[0]
        subject_id = otp_session[2]
        otp_year_id = otp_session[3]
        teacher_latitude = float(otp_session[4])
        teacher_longitude = float(otp_session[5])
        allowed_radius = int(otp_session[6])
        subject_name = otp_session[8]
        subject_dept_id = otp_session[9]
        year_name = otp_session[10]

        # Student same department ka hona chahiye
        if student_dept_id != subject_dept_id:
            mysql.connection.commit()

            return jsonify({
                "success": False,
                "message": "This OTP is not for your department."
            }), 403

        # Student same year ka hona chahiye
        if student_year_id != otp_year_id:
            mysql.connection.commit()

            return jsonify({
                "success": False,
                "message": "This OTP is not for your year."
            }), 403

        # Haversine formula se distance calculate karo
        earth_radius = 6371000

        teacher_lat_rad = math.radians(
            teacher_latitude
        )
        teacher_lon_rad = math.radians(
            teacher_longitude
        )
        student_lat_rad = math.radians(
            student_latitude
        )
        student_lon_rad = math.radians(
            student_longitude
        )

        latitude_difference = (
            student_lat_rad - teacher_lat_rad
        )

        longitude_difference = (
            student_lon_rad - teacher_lon_rad
        )

        haversine_value = (
            math.sin(latitude_difference / 2) ** 2
            + math.cos(teacher_lat_rad)
            * math.cos(student_lat_rad)
            * math.sin(longitude_difference / 2) ** 2
        )

        central_angle = 2 * math.atan2(
            math.sqrt(haversine_value),
            math.sqrt(1 - haversine_value)
        )

        distance_metres = earth_radius * central_angle

        if distance_metres > allowed_radius:
            mysql.connection.commit()

            return jsonify({
                "success": False,
                "message": (
                    "You are outside the allowed attendance area. "
                    f"Distance: {round(distance_metres)} metres. "
                    f"Allowed: {allowed_radius} metres."
                ),
                "distance": round(distance_metres, 2),
                "allowed_radius": allowed_radius
            }), 403

        attendance_date = datetime.now().date()

        # Present attendance insert/update
        cur.execute("""
            INSERT INTO attendance (
                student_id,
                subject_id,
                date,
                status
            )
            VALUES (%s, %s, %s, 'P')
            ON DUPLICATE KEY UPDATE
                status = 'P'
        """, (
            student_id,
            subject_id,
            attendance_date
        ))

        # Excel ke liye subject information
        cur.execute("""
            SELECT
                d.name,
                y.name,
                s.name
            FROM subjects s
            JOIN departments d
                ON d.id = s.dept_id
            JOIN years y
                ON y.id = s.year_id
            WHERE s.id = %s
        """, (subject_id,))

        subject_info = cur.fetchone()

        if not subject_info:
            mysql.connection.rollback()

            return jsonify({
                "success": False,
                "message": "Subject information not found."
            }), 404

        # Same subject/date ki attendance Excel ke liye
        cur.execute("""
            SELECT
                st.name,
                st.roll_no,
                a.status
            FROM attendance a
            JOIN students st
                ON st.id = a.student_id
            WHERE a.subject_id = %s
              AND a.date = %s
            ORDER BY st.roll_no
        """, (
            subject_id,
            attendance_date
        ))

        attendance_rows = cur.fetchall()

        mysql.connection.commit()

        department_name = subject_info[0]
        year_name = subject_info[1]
        subject_name = subject_info[2]

        update_attendance_excel(
            department_name,
            year_name,
            subject_name,
            attendance_date.strftime("%Y-%m-%d"),
            attendance_rows
        )

        return jsonify({
            "success": True,
            "message": (
                f"Attendance marked successfully for "
                f"{subject_name}."
            ),
            "subject": subject_name,
            "year": year_name,
            "status": "P",
            "distance": round(distance_metres, 2),
            "allowed_radius": allowed_radius,
            "otp_session_id": otp_session_id
        })

    except Exception as error:
        mysql.connection.rollback()

        print(
            "Student OTP attendance error:",
            error
        )

        return jsonify({
            "success": False,
            "message": (
                "Unable to mark attendance. "
                "Please try again."
            )
        }), 500

    finally:
        cur.close()

@app.route('/teacher/attendance/otp-status/<int:session_id>')
def teacher_otp_attendance_status(session_id):
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    cur = mysql.connection.cursor()

    try:
        # Logged-in teacher
        cur.execute("""
            SELECT id, dept_id
            FROM teachers
            WHERE user_id = %s
        """, (session['user_id'],))

        teacher = cur.fetchone()

        if not teacher:
            return jsonify({
                "success": False,
                "message": "Teacher not found."
            }), 404

        teacher_id = teacher[0]
        teacher_dept_id = teacher[1]

        # OTP session details
        cur.execute("""
            SELECT
                aos.subject_id,
                aos.year_id,
                DATE(aos.created_at),
                aos.is_active,
                aos.expires_at,
                s.name,
                y.name
            FROM attendance_otp_sessions aos
            JOIN subjects s
                ON s.id = aos.subject_id
            JOIN years y
                ON y.id = aos.year_id
            WHERE aos.id = %s
              AND aos.teacher_id = %s
            LIMIT 1
        """, (
            session_id,
            teacher_id
        ))

        otp_session = cur.fetchone()

        if not otp_session:
            return jsonify({
                "success": False,
                "message": "OTP session not found."
            }), 404

        subject_id = otp_session[0]
        year_id = otp_session[1]
        attendance_date = otp_session[2]
        is_active = bool(otp_session[3])
        expires_at = otp_session[4]
        subject_name = otp_session[5]
        year_name = otp_session[6]

        # Expired session inactive karo
        if expires_at <= datetime.now():
            cur.execute("""
                UPDATE attendance_otp_sessions
                SET is_active = 0
                WHERE id = %s
            """, (session_id,))

            mysql.connection.commit()
            is_active = False

        # Present/Absent count
        cur.execute("""
            SELECT
                SUM(
                    CASE
                        WHEN COALESCE(a.status, 'A') = 'P'
                        THEN 1
                        ELSE 0
                    END
                ) AS present_count,

                SUM(
                    CASE
                        WHEN COALESCE(a.status, 'A') = 'A'
                        THEN 1
                        ELSE 0
                    END
                ) AS absent_count,

                COUNT(st.id) AS total_count

            FROM students st

            LEFT JOIN attendance a
                ON a.student_id = st.id
                AND a.subject_id = %s
                AND a.date = %s

            WHERE st.dept_id = %s
              AND st.year_id = %s
        """, (
            subject_id,
            attendance_date,
            teacher_dept_id,
            year_id
        ))

        counts = cur.fetchone()

        present_count = int(counts[0] or 0)
        absent_count = int(counts[1] or 0)
        total_count = int(counts[2] or 0)

        return jsonify({
            "success": True,
            "session_id": session_id,
            "subject": subject_name,
            "year": year_name,
            "present": present_count,
            "absent": absent_count,
            "total": total_count,
            "is_active": is_active
        })

    except Exception as error:
        print("OTP status error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to load OTP attendance status."
        }), 500

    finally:
        cur.close()


@app.route('/teacher/attendance/submit', methods=['POST'])
def submit_teacher_attendance():
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Invalid request data."
        }), 400

    subject_id = data.get("subject_id")
    attendance_date = data.get("date")
    attendance = data.get("attendance", [])

    if not subject_id or not attendance_date or not attendance:
        return jsonify({
            "success": False,
            "message": "Missing attendance data."
        }), 400

    # Date format check
    try:
        selected_date = datetime.strptime(
            attendance_date,
            "%Y-%m-%d"
        ).date()
    except (ValueError, TypeError):
        return jsonify({
            "success": False,
            "message": "Invalid date format."
        }), 400

    # Future date block
    if selected_date > datetime.now().date():
        return jsonify({
            "success": False,
            "message": "Future date attendance is not allowed!"
        }), 400

    cur = mysql.connection.cursor()

    try:
        # Check logged-in teacher and subject assignment
        cur.execute("""
            SELECT t.id
            FROM teachers t
            JOIN teacher_subjects ts
                ON t.id = ts.teacher_id
            WHERE t.user_id = %s
              AND ts.subject_id = %s
        """, (session['user_id'], subject_id))

        teacher_subject = cur.fetchone()

        if not teacher_subject:
            return jsonify({
                "success": False,
                "message": "This subject is not assigned to you."
            }), 403

        present_count = 0
        absent_count = 0

        for item in attendance:
            student_id = item.get("student_id")
            status = item.get("status")

            if not student_id or status not in ("P", "A"):
                mysql.connection.rollback()

                return jsonify({
                    "success": False,
                    "message": "Invalid student attendance data."
                }), 400

            if status == "P":
                present_count += 1
            else:
                absent_count += 1

            cur.execute("""
                INSERT INTO attendance
                    (student_id, subject_id, date, status)
                VALUES
                    (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    status = VALUES(status)
            """, (
                student_id,
                subject_id,
                attendance_date,
                status
            ))

        # Subject information for Excel
        cur.execute("""
            SELECT
                d.name,
                y.name,
                s.name
            FROM subjects s
            JOIN departments d
                ON s.dept_id = d.id
            JOIN years y
                ON s.year_id = y.id
            WHERE s.id = %s
        """, (subject_id,))

        subject_info = cur.fetchone()

        if not subject_info:
            mysql.connection.rollback()

            return jsonify({
                "success": False,
                "message": "Subject information not found."
            }), 404

        # Attendance data for selected subject and date
        cur.execute("""
            SELECT
                st.name,
                st.roll_no,
                a.status
            FROM attendance a
            JOIN students st
                ON a.student_id = st.id
            WHERE a.subject_id = %s
              AND a.date = %s
            ORDER BY st.roll_no
        """, (subject_id, attendance_date))

        rows = cur.fetchall()

        mysql.connection.commit()

        department_name = subject_info[0]
        year_name = subject_info[1]
        subject_name = subject_info[2]

        update_attendance_excel(
            department_name,
            year_name,
            subject_name,
            attendance_date,
            rows
        )

        return jsonify({
            "success": True,
            "message": "Attendance submitted successfully!",
            "present": present_count,
            "absent": absent_count
        })

    except Exception as error:
        mysql.connection.rollback()
        print("Attendance submit error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to submit attendance. Please try again."
        }), 500

    finally:
        cur.close()

@app.route('/teacher/attendance')
def teacher_attendance():
    if not is_logged_in('teacher'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT t.name, t.dept_id, d.name
        FROM teachers t
        JOIN departments d ON t.dept_id = d.id
        WHERE t.user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()
    cur.close()

    return render_template(
        'teacher/attendance.html',
        teacher_name=teacher[0],
        dept_id=teacher[1],
        department=teacher[2]
    )

def update_attendance_excel(dept, year, subject, date, rows):
    base_folder = "attendance_excel"
    folder_path = os.path.join(base_folder, dept, year)
    os.makedirs(folder_path, exist_ok=True)

    safe_subject = subject.replace(" ", "_").replace("/", "_")
    file_path = os.path.join(folder_path, f"{safe_subject}.xlsx")

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name/Date", "Roll No", "Total", "%"])

    headers = [cell.value for cell in ws[1]]

    if "Total" not in headers:
        ws.cell(row=1, column=ws.max_column + 1).value = "Total"

    headers = [cell.value for cell in ws[1]]

    if "%" not in headers:
        ws.cell(row=1, column=ws.max_column + 1).value = "%"

    headers = [cell.value for cell in ws[1]]
    total_col = headers.index("Total") + 1

    if date in headers:
        date_col = headers.index(date) + 1
    else:
        ws.insert_cols(total_col)
        ws.cell(row=1, column=total_col).value = date

    headers = [cell.value for cell in ws[1]]
    total_col = headers.index("Total") + 1
    percent_col = headers.index("%") + 1
    date_col = headers.index(date) + 1

    student_rows = {}

    for r in range(2, ws.max_row + 1):
        roll = ws.cell(row=r, column=2).value
        if roll:
            student_rows[roll] = r

    for name, roll, status in rows:
        if roll in student_rows:
            row_num = student_rows[roll]
        else:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1).value = name
            ws.cell(row=row_num, column=2).value = roll

        ws.cell(row=row_num, column=date_col).value = status

    for r in range(2, ws.max_row + 1):
        present = 0
        total = 0

        for c in range(3, total_col):
            value = ws.cell(row=r, column=c).value

            if value in ["P", "A"]:
                total += 1
                if value == "P":
                    present += 1

        ws.cell(row=r, column=total_col).value = present

        if total > 0:
            ws.cell(row=r, column=percent_col).value = f"{round((present / total) * 100, 2)}%"
        else:
            ws.cell(row=r, column=percent_col).value = "0%"

    wb.save(file_path)
    
@app.route('/teacher/marks')
def teacher_marks():
    if not is_logged_in('teacher'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT t.name, d.name
        FROM teachers t
        JOIN departments d ON t.dept_id = d.id
        WHERE t.user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()
    cur.close()

    return render_template(
        'teacher/marks.html',
        teacher_name=teacher[0],
        department=teacher[1]
    )

@app.route('/teacher/marks/students/<year_name>/<subject_id>')
def teacher_marks_students(year_name, subject_id):
    if not is_logged_in('teacher'):
        return jsonify([])

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT dept_id
        FROM teachers
        WHERE user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()

    if not teacher:
        cur.close()
        return jsonify([])

    dept_id = teacher[0]

    cur.execute("""
        SELECT st.id, st.name, st.roll_no,
               IFNULL(m.internal_marks, 0),
               IFNULL(m.theory_marks, 0)
        FROM students st
        JOIN years y ON st.year_id = y.id
        LEFT JOIN marks m
            ON st.id = m.student_id
            AND m.subject_id = %s
        WHERE st.dept_id = %s
        AND y.name = %s
        ORDER BY st.roll_no
    """, (subject_id, dept_id, year_name))

    rows = cur.fetchall()
    cur.close()

    return jsonify([
        {
            "id": r[0],
            "name": r[1],
            "roll": r[2],
            "internal": r[3],
            "theory": r[4]
        }
        for r in rows
    ])

@app.route('/teacher/marks/submit', methods=['POST'])
def submit_teacher_marks():
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "message": "Invalid request data"
        }), 400

    subject_id = data.get("subject_id")
    marks = data.get("marks", [])

    if not subject_id or not marks:
        return jsonify({
            "success": False,
            "message": "Missing marks data"
        }), 400

    cur = mysql.connection.cursor()

    try:
        # Subject ka actual year nikalo
        cur.execute("""
            SELECT
                s.id,
                y.name
            FROM subjects s
            JOIN years y
                ON y.id = s.year_id
            WHERE s.id = %s
        """, (subject_id,))

        subject_info = cur.fetchone()

        if not subject_info:
            return jsonify({
                "success": False,
                "message": "Subject not found"
            }), 404

        year_name = subject_info[1]

        for item in marks:
            cur.execute("""
                INSERT INTO marks (
                    student_id,
                    subject_id,
                    internal_marks,
                    theory_marks,
                    total_marks,
                    year_name
                )
                VALUES (%s, %s, %s, %s, 100, %s)

                ON DUPLICATE KEY UPDATE
                    internal_marks = VALUES(internal_marks),
                    theory_marks = VALUES(theory_marks),
                    total_marks = 100,
                    year_name = VALUES(year_name)
            """, (
                item["student_id"],
                subject_id,
                item["internal"],
                item["theory"],
                year_name
            ))

        mysql.connection.commit()

        return jsonify({
            "success": True,
            "message": "Marks uploaded successfully!",
            "updated": len(marks),
            "year": year_name
        })

    except Exception as error:
        mysql.connection.rollback()

        print("Marks submit error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to upload marks."
        }), 500

    finally:
        cur.close()

@app.route('/teacher/defaulter')
def teacher_defaulter():
    if not is_logged_in('teacher'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT t.name, d.name
        FROM teachers t
        JOIN departments d ON t.dept_id = d.id
        WHERE t.user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()
    cur.close()

    return render_template(
        'teacher/defaulter.html',
        teacher_name=teacher[0],
        department=teacher[1]
    )

@app.route('/teacher/search/<keyword>')
def teacher_search_students(keyword):
    if not is_logged_in('teacher'):
        return jsonify({
            "success": False,
            "message": "Unauthorized"
        }), 401

    keyword = str(keyword).strip()

    if not keyword:
        return jsonify([])

    search_value = f"%{keyword}%"

    cur = mysql.connection.cursor()

    try:
        # Logged-in teacher ka department nikalo
        cur.execute("""
            SELECT
                id,
                dept_id
            FROM teachers
            WHERE user_id = %s
        """, (session['user_id'],))

        teacher = cur.fetchone()

        if not teacher:
            return jsonify({
                "success": False,
                "message": "Teacher not found"
            }), 404

        teacher_dept_id = teacher[1]

        # Sirf teacher ke department ke students search karo
        cur.execute("""
            SELECT
                st.id,
                st.name,
                st.roll_no,
                st.username,
                d.name,
                y.name,
                u.phone
            FROM students st

            JOIN departments d
                ON d.id = st.dept_id

            JOIN years y
                ON y.id = st.year_id

            JOIN users u
                ON u.id = st.user_id

            WHERE st.dept_id = %s
              AND (
                    st.name LIKE %s
                    OR st.roll_no LIKE %s
                    OR st.username LIKE %s
                    OR u.phone LIKE %s
                    OR y.name LIKE %s
                  )

            ORDER BY
                y.id ASC,
                st.roll_no ASC
        """, (
            teacher_dept_id,
            search_value,
            search_value,
            search_value,
            search_value,
            search_value
        ))

        rows = cur.fetchall()

        students = []

        for row in rows:
            students.append({
                "id": row[0],
                "name": row[1],
                "roll": row[2],
                "username": row[3],
                "department": row[4],
                "year": row[5],
                "phone": row[6] or "N/A"
            })

        return jsonify(students)

    except Exception as error:
        print("Teacher student search error:", error)

        return jsonify({
            "success": False,
            "message": "Unable to search students"
        }), 500

    finally:
        cur.close()

@app.route('/teacher/defaulters/<year_name>')
def teacher_defaulters_list(year_name):
    if not is_logged_in('teacher'):
        return jsonify([])

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT dept_id
        FROM teachers
        WHERE user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()

    if not teacher:
        cur.close()
        return jsonify([])

    dept_id = teacher[0]

    cur.execute("""
        SELECT 
            st.id,
            st.name,
            st.roll_no,
            y.name,
            u.phone,
            ROUND(
                (SUM(CASE WHEN a.status = 'P' THEN 1 ELSE 0 END) / COUNT(a.id)) * 100,
                2
            ) AS attendance_percent
        FROM students st
        JOIN users u ON st.user_id = u.id
        JOIN years y ON st.year_id = y.id
        JOIN attendance a ON st.id = a.student_id
        JOIN subjects s ON a.subject_id = s.id
        WHERE st.dept_id = %s
        AND y.name = %s
        AND s.year_id = y.id
        GROUP BY st.id, st.name, st.roll_no, y.name, u.phone
        HAVING attendance_percent < 75
        ORDER BY attendance_percent ASC
    """, (dept_id, year_name))

    rows = cur.fetchall()
    cur.close()

    return jsonify([
        {
            "id": r[0],
            "name": r[1],
            "roll": r[2],
            "year": r[3],
            "phone": r[4],
            "percent": r[5]
        }
        for r in rows
    ])

@app.route('/teacher/profile')
def teacher_profile():
    if not is_logged_in('teacher'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT t.name, t.username, u.password, u.phone, d.name
        FROM teachers t
        JOIN users u ON t.user_id = u.id
        JOIN departments d ON t.dept_id = d.id
        WHERE t.user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()
    cur.close()

    return render_template(
        'teacher/profile.html',
        name=teacher[0],
        username=teacher[1],
        password=teacher[2],
        phone=teacher[3],
        department=teacher[4]
    )

@app.route('/teacher/change-password', methods=['GET', 'POST'])
def teacher_change_password():
    if not is_logged_in('teacher'):
        return redirect(url_for('login'))

    if request.method == 'POST':

        current_password = request.form['current_password']
        new_password = request.form['new_password']

        cur = mysql.connection.cursor()

        cur.execute("""
            SELECT password
            FROM users
            WHERE id = %s
        """, (session['user_id'],))

        user = cur.fetchone()

        if not user:
            cur.close()
            flash("User not found.", "error")
            return redirect(url_for('teacher_change_password'))

        if user[0] != current_password:
            cur.close()
            flash("Old password is incorrect.", "error")
            return redirect(url_for('teacher_change_password'))

        if current_password == new_password:
            cur.close()
            flash("New password cannot be same as old password.", "error")
            return redirect(url_for('teacher_change_password'))

        cur.execute("""
            UPDATE users
            SET password = %s
            WHERE id = %s
        """, (new_password, session['user_id']))

        mysql.connection.commit()
        cur.close()

        flash("Password changed successfully.", "success")
        return redirect(url_for('teacher_change_password'))

    return render_template("teacher/change_password.html")  

@app.route('/teacher/students/<year_name>')
def teacher_students(year_name):
    if not is_logged_in('teacher'):
        return jsonify([])

    cur = mysql.connection.cursor()

    # Teacher ka department
    cur.execute("""
        SELECT dept_id
        FROM teachers
        WHERE user_id = %s
    """, (session['user_id'],))

    teacher = cur.fetchone()

    if not teacher:
        cur.close()
        return jsonify([])

    dept_id = teacher[0]

    cur.execute("""
        SELECT st.id,
               st.name,
               st.roll_no,
               st.username,
               u.phone
        FROM students st
        JOIN users u ON st.user_id = u.id
        JOIN years y ON st.year_id = y.id
        WHERE st.dept_id = %s
          AND y.name = %s
        ORDER BY st.roll_no
    """, (dept_id, year_name))

    rows = cur.fetchall()
    cur.close()

    students = []

    for r in rows:
        students.append({
            "id": r[0],
            "name": r[1],
            "roll": r[2],
            "username": r[3],
            "phone": r[4]
        })

    return jsonify(students)


# =========================
# STUDENT ROUTES
# =========================
@app.route('/student/dashboard')
def student_dashboard():
    if not is_logged_in('student'):
        return redirect(url_for('login'))
    return redirect(url_for('student_attendance'))


@app.route('/student/attendance')
def student_attendance():
    if not is_logged_in('student'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()

    try:
        cur.execute("""
            SELECT
                st.id,
                st.name,
                d.name,
                y.name
            FROM students st
            JOIN departments d
                ON st.dept_id = d.id
            JOIN years y
                ON st.year_id = y.id
            WHERE st.user_id = %s
        """, (session['user_id'],))

        student = cur.fetchone()

        if not student:
            flash(
                'Student profile not found!',
                'error'
            )

            return redirect(url_for('login'))

        student_id = student[0]
        student_name = student[1]
        department = student[2]
        year = student[3]

        current_summary = get_attendance_summary(
            student_id,
            year
        )

        fy_summary = get_attendance_summary(
            student_id,
            'FY'
        )

        sy_summary = get_attendance_summary(
            student_id,
            'SY'
        )

        ty_summary = get_attendance_summary(
            student_id,
            'TY'
        )

        total = current_summary['total']
        present = current_summary['present']
        absent = current_summary['absent']
        leave = current_summary['leave']
        attendance_percent = current_summary['percent']

        if total > 0:
            present_degree = round(
                (present / total) * 360,
                2
            )

            absent_degree = round(
                (
                    (present + absent) / total
                ) * 360,
                2
            )
        else:
            present_degree = 0
            absent_degree = 0

        # Student attendance history
        cur.execute("""
            SELECT
                a.date,
                s.name,
                y.name,
                a.status
            FROM attendance a
            JOIN subjects s
                ON s.id = a.subject_id
            JOIN years y
                ON y.id = s.year_id
            WHERE a.student_id = %s
            ORDER BY
                a.date DESC
                LIMIT 10
        """, (student_id,))

        history_rows = cur.fetchall()

        attendance_history = [
            {
                "date": row[0].strftime("%d-%m-%Y")
                if row[0]
                else "-",

                "subject": row[1] or "-",

                "year": row[2] or "-",

                "status": row[3] or "-"
            }
            for row in history_rows
        ]

        return render_template(
            'student/attendance.html',

            student_name=student_name,
            department=department,
            year=year,

            present=present,
            absent=absent,
            leave=leave,
            total=total,

            attendance_percent=attendance_percent,
            present_degree=present_degree,
            absent_degree=absent_degree,

            fy_percent=fy_summary['percent'],
            sy_percent=sy_summary['percent'],
            ty_percent=ty_summary['percent'],

            attendance_history=attendance_history
        )

    except Exception as error:
        print(
            "Student attendance page error:",
            error
        )

        flash(
            "Unable to load attendance details.",
            "error"
        )

        return redirect(url_for('student_profile'))

    finally:
        cur.close()


@app.route('/student/result')
def student_result():
    if not is_logged_in('student'):
        return redirect(url_for('login'))

    selected_year = request.args.get('year', 'FY')

    cur = mysql.connection.cursor()

    cur.execute("""
        SELECT st.id, st.name, st.roll_no, d.name, y.name
        FROM students st
        JOIN departments d ON st.dept_id = d.id
        JOIN years y ON st.year_id = y.id
        WHERE st.user_id = %s
    """, (session['user_id'],))

    student = cur.fetchone()

    cur.execute("""
        SELECT
            s.name,
            m.internal_marks,
            m.theory_marks,
            m.total_marks
        FROM marks m
        JOIN subjects s
        ON m.subject_id = s.id
        JOIN students st
        ON st.id = m.student_id
        JOIN years y
        ON y.id = s.year_id
        WHERE
        m.student_id = %s
        AND y.name = %s
        ORDER BY s.id
    """, (student[0],selected_year))

    rows = cur.fetchall()
    cur.close()

    results = []
    obtained_total = 0
    max_total = 0
    fail = False

    for r in rows:
        subject_total = r[1] + r[2]
        percent = round((subject_total / r[3]) * 100, 2)

        if percent < 40:
            fail = True

        obtained_total += subject_total
        max_total += r[3]

        results.append({
            'subject': r[0],
            'internal': r[1],
            'theory': r[2],
            'total': subject_total,
            'max': r[3],
            'percent': percent
        })

    if max_total > 0:
        overall_percent = round((obtained_total / max_total) * 100, 2)
        sgpa = round(overall_percent / 10, 2)
        cgpa = sgpa
        status = "FAIL" if fail else "PASS"
    else:
        overall_percent = 0
        sgpa = 0
        cgpa = 0
        status = "NOT AVAILABLE"

    return render_template(
        'student/result.html',
        student=student,
        selected_year=selected_year,
        results=results,
        obtained_total=obtained_total,
        max_total=max_total,
        overall_percent=overall_percent,
        sgpa=sgpa,
        cgpa=cgpa,
        status=status
    )


@app.route('/student/profile')
def student_profile():
    if not is_logged_in('student'):
        return redirect(url_for('login'))

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT st.name, st.roll_no, st.username, u.password, u.phone, d.name, y.name
        FROM students st
        JOIN users u ON st.user_id = u.id
        JOIN departments d ON st.dept_id = d.id
        JOIN years y ON st.year_id = y.id
        WHERE st.user_id = %s
    """, (session['user_id'],))

    student = cur.fetchone()
    cur.close()

    if not student:
        flash('Student profile not found!', 'error')
        return redirect(url_for('student_attendance'))

    hour = datetime.now().hour

    if hour < 12:
        greeting = "Good Morning"
    elif hour < 17:
        greeting = "Good Afternoon"
    else:
        greeting = "Good Evening"

    return render_template(
        'student/profile.html',
        greeting=greeting,
        name=student[0],
        roll_no=student[1],
        username=student[2],
        password=student[3],
        phone=student[4],
        department=student[5],
        year=student[6]
    )


@app.route('/student/change-password', methods=['GET', 'POST'])
def student_change_password():
    if not is_logged_in('student'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        old_password = request.form['old_password']
        new_password = request.form['new_password']

        cur = mysql.connection.cursor()
        cur.execute("SELECT password FROM users WHERE id = %s", (session['user_id'],))
        user = cur.fetchone()

        if user and str(user[0]) == str(old_password):
            cur.execute(
                "UPDATE users SET password = %s WHERE id = %s",
                (new_password, session['user_id'])
            )
            mysql.connection.commit()
            flash('Password changed successfully!', 'success')
        else:
            flash('Old password is incorrect!', 'error')

        cur.close()
        return redirect(url_for('student_change_password'))

    return render_template('student/password.html')


# =========================
# LOGOUT
# =========================
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# =========================
# RUN
# =========================
if __name__ == '__main__':
    app.run(debug=True)
