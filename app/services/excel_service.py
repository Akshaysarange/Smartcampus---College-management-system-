import os

from openpyxl import Workbook, load_workbook

BASE_FOLDER = "attendance_excel"


def update_attendance_excel(dept, year, subject, date, rows):
    """Write daily attendance into a per-subject Excel workbook."""
    folder_path = os.path.join(BASE_FOLDER, dept, year)
    os.makedirs(folder_path, exist_ok=True)

    safe_subject = subject.replace(" ", "_").replace("/", "_")
    file_path = os.path.join(folder_path, f"{safe_subject}.xlsx")

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name/Date", "Roll No", "Total", "%"])

    headers = [cell.value for cell in ws[1]]

    if "Total" not in headers:
        ws.cell(row=1, column=ws.max_column + 1).value = "Total"

    headers = [cell.value for cell in ws[1]]

    if "%" not in headers:
        ws.cell(row=1, column=ws.max_column + 1).value = "%"

    headers = [cell.value for cell in ws[1]]
    total_col = headers.index("Total") + 1

    if date in headers:
        date_col = headers.index(date) + 1
    else:
        ws.insert_cols(total_col)
        ws.cell(row=1, column=total_col).value = date

    headers = [cell.value for cell in ws[1]]
    total_col = headers.index("Total") + 1
    percent_col = headers.index("%") + 1
    date_col = headers.index(date) + 1

    student_rows = {}

    for r in range(2, ws.max_row + 1):
        roll = ws.cell(row=r, column=2).value
        if roll:
            student_rows[roll] = r

    for name, roll, status in rows:
        if roll in student_rows:
            row_num = student_rows[roll]
        else:
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1).value = name
            ws.cell(row=row_num, column=2).value = roll

        ws.cell(row=row_num, column=date_col).value = status

    for r in range(2, ws.max_row + 1):
        present = 0
        total = 0

        for c in range(3, total_col):
            value = ws.cell(row=r, column=c).value

            if value in ("P", "A"):
                total += 1
                if value == "P":
                    present += 1

        ws.cell(row=r, column=total_col).value = present

        if total > 0:
            ws.cell(
                row=r,
                column=percent_col,
            ).value = f"{round((present / total) * 100, 2)}%"
        else:
            ws.cell(row=r, column=percent_col).value = "0%"

    wb.save(file_path)
